"""
app.py — TrackED Web Application
"""
import os, sys, json, shutil, re, sqlite3
from openpyxl import load_workbook
from datetime import datetime
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, send_file, jsonify, g)

sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'scripts'))

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-change-in-prod')
app.jinja_env.globals['enumerate'] = enumerate

DATA_DIR     = os.environ.get('DATA_DIR', '/tmp/tracked')
DB_PATH      = os.path.join(DATA_DIR, 'tracked.db')
PROGRAMS_DIR = os.path.join(DATA_DIR, 'programs')
UPLOADS_DIR  = os.path.join(DATA_DIR, 'uploads')

for _d in [DATA_DIR, PROGRAMS_DIR, UPLOADS_DIR]:
    os.makedirs(_d, exist_ok=True)


# ── Feedback paste processor ────────────────────────────────────────────────

def process_feedback_paste(pasted_text, module_name, mentor_name,
                            log_path, report_path):
    """
    Parse pasted feedback text and append to report.
    Returns {'success': True, 'new_rows': N, 'skipped_rows': N}
    """
    if not pasted_text or not pasted_text.strip():
        return {'success': False, 'error': 'No text pasted.'}
    if not module_name or not module_name.strip():
        return {'success': False, 'error': 'Module name is required.'}

    lines = [l for l in pasted_text.strip().splitlines() if l.strip()]
    if not lines:
        return {'success': False, 'error': 'No data found in pasted text.'}

    # Detect separator
    first_line = lines[0]
    sep = '\t' if '\t' in first_line else ','

    # Parse header row (first line)
    headers = [h.strip().strip('"') for h in first_line.split(sep)]

    # Find key columns
    def find_col(candidates):
        for cand in candidates:
            for i, h in enumerate(headers):
                if cand.lower() in h.lower():
                    return i
        return None

    participant_col = find_col(['participant name', 'participant', 'name', 'respondent', 'your name', 'employee name', 'full name', 'username'])
    rating_col      = find_col(['rating', 'rate', 'score', 'stars'])
    takeaway_col    = find_col(['takeaway', 'key learning', 'what did you'])
    specific_col    = find_col(['specific', 'feedback for', 'delivery', 'mentor'])
    other_col       = find_col(['other', 'additional', 'anything'])
    date_col        = find_col(['start time', 'timestamp', 'date', 'completion'])

    if participant_col is None:
        # Try to find any column that looks like names (not numbers/dates)
        # Default to column 0 but warn
        participant_col = 0

    # Load workbook
    if not os.path.exists(report_path):
        return {'success': False, 'error': 'Report not found. Upload attendance first.'}

    wb = load_workbook(report_path)
    if 'Feedback' not in wb.sheetnames:
        return {'success': False, 'error': 'No Feedback sheet in report.'}

    ws = wb['Feedback']

    # Find next empty row + max Sno
    next_row = 2
    while ws.cell(next_row, 1).value is not None:
        next_row += 1

    max_sno = 0
    for row in range(2, next_row):
        v = ws.cell(row, 1).value
        try:
            if v: max_sno = max(max_sno, int(str(v)))
        except: pass

    # Load existing keys for dedup
    log = {}
    if os.path.exists(log_path):
        with open(log_path) as f:
            try: log = json.load(f)
            except: pass
    done_keys  = set(log.get('processed_feedback_keys', []))
    new_count     = 0
    skip_count    = 0
    skipped_names = []
    timestamp  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def get_cell(row_cells, idx):
        if idx is None or idx >= len(row_cells):
            return ''
        return row_cells[idx].strip().strip('"')

    # Skip header line, process data lines
    data_lines = lines[1:] if len(lines) > 1 else lines

    for line in data_lines:
        cells = line.split(sep)
        participant = get_cell(cells, participant_col)
        if not participant:
            continue

        rating      = get_cell(cells, rating_col)
        takeaways   = get_cell(cells, takeaway_col)
        specific    = get_cell(cells, specific_col)
        other       = get_cell(cells, other_col)
        ts          = get_cell(cells, date_col) or timestamp

        # Dedup key: participant + module + takeaways content
        # Using content-based key so:
        # - Same feedback re-pasted = skipped (same participant+module+takeaway)
        # - Same person, different feedback same day = allowed (different takeaway)
        takeaway_key = takeaways[:50].lower().strip() if takeaways else ''
        rating_key   = str(rating).strip()
        key = f"{participant.lower().strip()}|{module_name.lower()}|{takeaway_key}|{rating_key}"
        if key in done_keys:
            skip_count += 1
            skipped_names.append(participant)
            continue

        max_sno += 1
        ws.cell(next_row, 1, max_sno)
        ws.cell(next_row, 2, ts)
        ws.cell(next_row, 3, module_name)
        ws.cell(next_row, 4, mentor_name)
        ws.cell(next_row, 5, participant)
        ws.cell(next_row, 6, takeaways)
        try:    ws.cell(next_row, 7, float(rating))
        except: ws.cell(next_row, 7, rating)
        ws.cell(next_row, 8, specific)
        ws.cell(next_row, 9, other)

        done_keys.add(key)
        next_row  += 1
        new_count += 1

    wb.save(report_path)

    log['processed_feedback_keys'] = list(done_keys)
    log.setdefault('feedback_runs', []).append({
        'method':       'paste',
        'module':       module_name,
        'mentor':       mentor_name,
        'new_rows':     new_count,
        'skipped':      skip_count,
        'processed_at': datetime.now().isoformat()
    })
    os.makedirs(os.path.dirname(log_path) or '.', exist_ok=True)
    with open(log_path, 'w') as f:
        json.dump(log, f, indent=2, default=str)

    return {'success': True, 'new_rows': new_count, 'skipped_rows': skip_count, 'skipped_details': skipped_names}


# ── Database ──────────────────────────────────────────────────────────────────

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH, detect_types=sqlite3.PARSE_DECLTYPES)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db:
        db.close()

def init_db():
    db = sqlite3.connect(DB_PATH)
    db.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL, name TEXT NOT NULL,
            password TEXT NOT NULL, is_admin INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS programs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slug TEXT UNIQUE NOT NULL, name TEXT NOT NULL,
            client_name TEXT NOT NULL, settings TEXT NOT NULL,
            created_by INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS program_access (
            program_id INTEGER NOT NULL, user_id INTEGER NOT NULL,
            PRIMARY KEY (program_id, user_id)
        );
    ''')
    db.commit()
    admin_email    = os.environ.get('ADMIN_EMAIL',    'admin@tracked.app')
    admin_password = os.environ.get('ADMIN_PASSWORD', 'admin123')
    admin_name     = os.environ.get('ADMIN_NAME',     'Admin')
    existing = db.execute('SELECT id FROM users WHERE email=?',
                          (admin_email,)).fetchone()
    if not existing:
        db.execute(
            'INSERT INTO users (email, name, password, is_admin) VALUES (?,?,?,1)',
            (admin_email, admin_name, generate_password_hash(admin_password))
        )
        db.commit()
        print(f'✅ Admin account created: {admin_email}')
    db.close()


# ── Helpers ───────────────────────────────────────────────────────────────────

def slugify(text):
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')

def program_dir(slug):
    return os.path.join(PROGRAMS_DIR, slug)

def report_path(slug):
    return os.path.join(program_dir(slug), 'output', 'Master_Report.xlsx')

def log_path(slug):
    return os.path.join(program_dir(slug), 'logs', 'run_log.json')

def load_log(slug):
    path = log_path(slug)
    if os.path.exists(path):
        with open(path) as f:
            try: return json.load(f)
            except: pass
    return {'runs': []}

def save_log(slug, log):
    path = log_path(slug)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w') as f:
        json.dump(log, f, indent=2, default=str)

def ensure_dirs(slug):
    for sub in ['output', 'logs', 'uploads']:
        os.makedirs(os.path.join(program_dir(slug), sub), exist_ok=True)

def get_program(slug, user_id):
    return get_db().execute('''
        SELECT p.* FROM programs p
        JOIN program_access pa ON pa.program_id = p.id
        WHERE p.slug=? AND pa.user_id=?
    ''', (slug, user_id)).fetchone()

def get_user_programs(user_id):
    return get_db().execute('''
        SELECT p.* FROM programs p
        JOIN program_access pa ON pa.program_id = p.id
        WHERE pa.user_id=? ORDER BY p.created_at DESC
    ''', (user_id,)).fetchall()

def parse_names(raw):
    return [n.strip() for n in re.split(r'[,\n]', raw) if n.strip()]


# ── Auth ──────────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in.', 'info')
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if not session.get('is_admin'):
            flash('Admin access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        pw    = request.form.get('password', '').strip()
        user  = get_db().execute('SELECT * FROM users WHERE email=?', (email,)).fetchone()
        if user and check_password_hash(user['password'], pw):
            session.clear()
            session['user_id']  = user['id']
            session['email']    = user['email']
            session['name']     = user['name']
            session['is_admin'] = bool(user['is_admin'])
            return redirect(request.args.get('next', url_for('dashboard')))
        flash('Incorrect email or password.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email   = request.form.get('email', '').strip().lower()
        new_pw  = request.form.get('new_password', '').strip()
        confirm = request.form.get('confirm_password', '').strip()
        token   = request.form.get('token', '').strip()
        if new_pw != confirm:
            flash('Passwords do not match.', 'error')
            return render_template('forgot_password.html')
        if len(new_pw) < 6:
            flash('Password must be at least 6 characters.', 'error')
            return render_template('forgot_password.html')
        reset_file = os.path.join(DATA_DIR, 'reset_requests.json')
        resets = {}
        if os.path.exists(reset_file):
            with open(reset_file) as f:
                try: resets = json.load(f)
                except: pass
        if email not in resets or resets[email].get('token') != token:
            flash('Invalid or expired link.', 'error')
            return render_template('forgot_password.html')
        user = get_db().execute('SELECT id FROM users WHERE email=?', (email,)).fetchone()
        if not user:
            flash('No account found.', 'error')
            return render_template('forgot_password.html')
        db = get_db()
        db.execute('UPDATE users SET password=? WHERE id=?',
                   (generate_password_hash(new_pw), user['id']))
        db.commit()
        del resets[email]
        with open(reset_file, 'w') as f: json.dump(resets, f)
        flash('Password reset. Please log in.', 'success')
        return redirect(url_for('login'))
    return render_template('forgot_password.html',
                           token=request.args.get('token',''),
                           email=request.args.get('email',''))

@app.route('/dashboard')
@login_required
def dashboard():
    rows     = get_user_programs(session['user_id'])
    programs = []
    for p in rows:
        log = load_log(p['slug'])
        settings_p  = json.loads(p['settings'])
        total_done  = len(log.get('runs', []))
        # Status: Completed if report exists and has sessions, else In Progress
        if total_done == 0:
            status = 'not_started'
        else:
            status = 'completed' if log.get('completed') else 'in_progress'
        programs.append({
            'slug':          p['slug'],
            'name':          p['name'],
            'client_name':   p['client_name'],
            'sessions_done': total_done,
            'report_exists': os.path.exists(report_path(p['slug'])),
            'last_updated':  log['runs'][-1]['processed'][:10]
                             if log.get('runs') else '—',
            'status':        status,
        })
    return render_template('dashboard.html', programs=programs)

@app.route('/programs/new', methods=['GET', 'POST'])
@login_required
def new_program():
    if request.method == 'POST':
        name   = request.form.get('name', '').strip()
        client = request.form.get('client', '').strip()
        if not name or not client:
            flash('Name and client are required.', 'error')
            return render_template('new_program.html')
        settings = {
            'name_format':   request.form.get('name_format', 'auto'),
            'exclude_names': parse_names(request.form.get('exclude_names', '')),
            'mentor_names':  parse_names(request.form.get('mentor_names', '')),
            'threshold':     int(request.form.get('threshold', 50)) / 100,
        }
        slug = slugify(f"{client}-{name}")
        db   = get_db()
        if db.execute('SELECT id FROM programs WHERE slug=?', (slug,)).fetchone():
            slug = f"{slug}-{datetime.now().strftime('%m%d%H%M')}"
        db.execute(
            'INSERT INTO programs (slug,name,client_name,settings,created_by) VALUES (?,?,?,?,?)',
            (slug, name, client, json.dumps(settings), session['user_id'])
        )
        db.commit()
        prog = db.execute('SELECT id FROM programs WHERE slug=?', (slug,)).fetchone()
        db.execute('INSERT INTO program_access (program_id,user_id) VALUES (?,?)',
                   (prog['id'], session['user_id']))
        db.commit()
        ensure_dirs(slug)
        flash(f'Program "{name}" created!', 'success')
        return redirect(url_for('program_detail', slug=slug))
    return render_template('new_program.html')

@app.route('/programs/<slug>')
@login_required
def program_detail(slug):
    p = get_program(slug, session['user_id'])
    if not p:
        flash('Program not found or no access.', 'error')
        return redirect(url_for('dashboard'))
    return render_template('program.html',
                           program=p,
                           settings=json.loads(p['settings']),
                           log=load_log(slug),
                           report_exists=os.path.exists(report_path(slug)))

@app.route('/programs/<slug>/upload-attendance', methods=['POST'])
@login_required
def upload_attendance(slug):
    p = get_program(slug, session['user_id'])
    if not p: return jsonify({'success': False, 'error': 'Not found'}), 404
    file = request.files.get('csv_file')
    if not file or not file.filename.endswith('.csv'):
        return jsonify({'success': False, 'error': 'Please upload a .csv file'})
    ensure_dirs(slug)
    upload_path = os.path.join(program_dir(slug), 'uploads',
                               f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
    file.save(upload_path)
    try:
        from smart_report import process_csv
        s = json.loads(p['settings'])
        result = process_csv(upload_path, report_path(slug),
                             s.get('exclude_names',[]), s.get('mentor_names',[]),
                             s.get('threshold',0.5), s.get('name_format','auto'))
        if not result['success']:
            return jsonify({'success': False, 'error': result['error']})
        log = load_log(slug)
        log.setdefault('runs', []).append({
            'file': file.filename, 'date': result['date'],
            'title': result['title'], 'session_num': result['session_num'],
            'present': result['present'], 'absent': result['absent'],
            'processed': datetime.now().isoformat()
        })
        save_log(slug, log)
        msg = (f"Session {result['session_num']} — {result['present']} present"
               + (f", {len(result['absent'])} absent: {', '.join(result['absent'])}"
                  if result['absent'] else ', full attendance!'))
        return jsonify({'success': True, 'message': msg,
                        'present': result['present'], 'absent': result['absent'],
                        'session_num': result['session_num'], 'date': result['date'],
                        'title': result['title']})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/programs/<slug>/upload-feedback', methods=['POST'])
@login_required
def upload_feedback(slug):
    p = get_program(slug, session['user_id'])
    if not p: return jsonify({'success': False, 'error': 'Not found'}), 404
    file = request.files.get('feedback_file')
    if not file or not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'success': False, 'error': 'Please upload an .xlsx file'})
    if not os.path.exists(report_path(slug)):
        return jsonify({'success': False, 'error': 'Process attendance first.'})
    ensure_dirs(slug)
    upload_path = os.path.join(program_dir(slug), 'uploads',
                               f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
    file.save(upload_path)
    config_path = os.path.join(program_dir(slug), 'config.json')
    if not os.path.exists(config_path):
        with open(config_path, 'w') as f:
            json.dump({'output_file': report_path(slug), 'feedback_column_map': {}}, f)
    try:
        from process_feedback import process_feedback
        result = process_feedback(upload_path, config_path,
                                  log_path(slug), report_path(slug),
                                  prog_dir=program_dir(slug))
        if not result or not result.get('success'):
            err  = result.get('error','Unknown') if result else 'Unknown'
            cols = result.get('available_columns',[]) if result else []
            nm   = result.get('needs_mapping',False) if result else False
            if nm and result.get('auto_map'):
                lg = load_log(slug)
                lg['feedback_columns'] = {k:v for k,v in result['auto_map'].items() if v}
                save_log(slug, lg)
            return jsonify({'success': False, 'error': err,
                            'needs_mapping': nm, 'available_columns': cols,
                            'mapping_url': url_for('feedback_columns', slug=slug) if nm else ''})
        return jsonify({'success': True,
                        'new_rows': result['new_rows'],
                        'skipped':  result['skipped_rows'],
                        'message':  f"{result['new_rows']} new rows added"
                                    + (f", {result['skipped_rows']} skipped"
                                       if result['skipped_rows'] else '')})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/programs/<slug>/download')
@login_required
def download_report(slug):
    p = get_program(slug, session['user_id'])
    if not p:
        flash('Not found.', 'error')
        return redirect(url_for('dashboard'))
    rpath = report_path(slug)
    if not os.path.exists(rpath):
        flash('No report yet.', 'error')
        return redirect(url_for('program_detail', slug=slug))
    log   = load_log(slug)
    today = datetime.now().strftime('%d%b%Y')
    fname = (f"{p['client_name']}_{p['name']}_"
             f"{len(log.get('runs',[]))}Sessions_{today}.xlsx").replace(' ','_')
    return send_file(rpath, as_attachment=True, download_name=fname)

@app.route('/programs/<slug>/edit', methods=['GET', 'POST'])
@login_required
def edit_program(slug):
    p = get_program(slug, session['user_id'])
    if not p:
        flash('Not found.', 'error')
        return redirect(url_for('dashboard'))
    settings = json.loads(p['settings'])
    if request.method == 'POST':
        new_s = {
            'name_format':   request.form.get('name_format','auto'),
            'exclude_names': parse_names(request.form.get('exclude_names','')),
            'mentor_names':  parse_names(request.form.get('mentor_names','')),
            'threshold':     int(request.form.get('threshold',50)) / 100,
        }
        db = get_db()
        db.execute('UPDATE programs SET name=?,client_name=?,settings=? WHERE slug=?',
                   (request.form.get('name'), request.form.get('client'),
                    json.dumps(new_s), slug))
        db.commit()
        flash('Program updated.', 'success')
        return redirect(url_for('program_detail', slug=slug))
    return render_template('edit_program.html', program=p, settings=settings)

@app.route('/programs/<slug>/delete', methods=['POST'])
@login_required
def delete_program(slug):
    p = get_program(slug, session['user_id'])
    if not p:
        flash('Not found.', 'error')
        return redirect(url_for('dashboard'))
    db = get_db()
    db.execute('DELETE FROM program_access WHERE program_id=?', (p['id'],))
    db.execute('DELETE FROM programs WHERE id=?', (p['id'],))
    db.commit()
    pdir = program_dir(slug)
    if os.path.exists(pdir): shutil.rmtree(pdir)
    flash('Program deleted.', 'success')
    return redirect(url_for('dashboard'))

@app.route('/programs/<slug>/undo-session', methods=['POST'])
@login_required
def undo_session(slug):
    p = get_program(slug, session['user_id'])
    if not p: return jsonify({'success': False, 'error': 'Not found'}), 404
    sn  = request.form.get('session_num', type=int)
    log = load_log(slug)
    log['runs'] = [r for r in log.get('runs',[]) if r.get('session_num') != sn]
    save_log(slug, log)
    rpath = report_path(slug)
    if os.path.exists(rpath): os.remove(rpath)
    remaining = sorted(log['runs'], key=lambda r: r.get('session_num',0))
    if remaining:
        from smart_report import process_csv
        s  = json.loads(p['settings'])
        uf = os.path.join(program_dir(slug), 'uploads')
        for run in remaining:
            cands = [f for f in os.listdir(uf) if run['file'] in f] if os.path.exists(uf) else []
            if cands:
                process_csv(os.path.join(uf, cands[0]), rpath,
                            s.get('exclude_names',[]), s.get('mentor_names',[]),
                            s.get('threshold',0.5), s.get('name_format','auto'))
    flash(f'Session {sn} removed.', 'success')
    return redirect(url_for('program_detail', slug=slug))

@app.route('/programs/<slug>/rebuild', methods=['POST'])
@login_required
def rebuild_report(slug):
    p = get_program(slug, session['user_id'])
    if not p:
        flash('Not found.', 'error')
        return redirect(url_for('dashboard'))
    rpath = report_path(slug)
    uf    = os.path.join(program_dir(slug), 'uploads')
    log   = load_log(slug)
    runs  = sorted(log.get('runs',[]), key=lambda r: r.get('session_num',0))
    if not runs:
        flash('No sessions to rebuild from.', 'error')
        return redirect(url_for('program_detail', slug=slug))
    if os.path.exists(rpath): os.remove(rpath)
    from smart_report import process_csv
    s = json.loads(p['settings'])
    rebuilt, failed = 0, []
    for run in runs:
        cands = [f for f in os.listdir(uf) if run['file'] in f] if os.path.exists(uf) else []
        if cands:
            res = process_csv(os.path.join(uf, cands[0]), rpath,
                              s.get('exclude_names',[]), s.get('mentor_names',[]),
                              s.get('threshold',0.5), s.get('name_format','auto'))
            if res['success']: rebuilt += 1
            else: failed.append(run['file'])
        else: failed.append(run['file'])
    if failed:
        flash(f'Rebuilt {rebuilt} sessions. Missing: {", ".join(failed)}. Re-upload them.', 'info')
    else:
        flash(f'Report rebuilt from {rebuilt} sessions.', 'success')
    return redirect(url_for('program_detail', slug=slug))

@app.route('/programs/<slug>/share', methods=['POST'])
@login_required
def share_program(slug):
    p = get_program(slug, session['user_id'])
    if not p:
        flash('Not found.', 'error')
        return redirect(url_for('dashboard'))
    email = request.form.get('email','').strip().lower()
    db    = get_db()
    user  = db.execute('SELECT id FROM users WHERE email=?', (email,)).fetchone()
    if not user:
        flash(f'No account found for {email}', 'error')
        return redirect(url_for('program_detail', slug=slug))
    try:
        db.execute('INSERT INTO program_access (program_id,user_id) VALUES (?,?)',
                   (p['id'], user['id']))
        db.commit()
        flash(f'Shared with {email}', 'success')
    except sqlite3.IntegrityError:
        flash(f'{email} already has access.', 'info')
    return redirect(url_for('program_detail', slug=slug))

@app.route('/programs/<slug>/feedback-columns', methods=['GET', 'POST'])
@login_required
def feedback_columns(slug):
    p = get_program(slug, session['user_id'])
    if not p:
        flash('Not found.', 'error')
        return redirect(url_for('dashboard'))
    lf = log_path(slug)
    ld = {}
    if os.path.exists(lf):
        with open(lf) as f:
            try: ld = json.load(f)
            except: pass
    cm = ld.get('feedback_columns', {})
    if request.method == 'POST':
        nm = {k: request.form.get(k,'').strip()
              for k in ['date','participant','takeaways','rating','specific','other']
              if request.form.get(k,'').strip()}
        ld['feedback_columns'] = nm
        os.makedirs(os.path.dirname(lf), exist_ok=True)
        with open(lf,'w') as f: json.dump(ld, f, indent=2)
        flash('Column mapping saved.', 'success')
        return redirect(url_for('program_detail', slug=slug))
    return render_template('feedback_columns.html', program=p, current_map=cm)

@app.route('/programs/<slug>/edit-report')
@login_required
def edit_report(slug):
    p = get_program(slug, session['user_id'])
    if not p:
        flash('Not found.', 'error')
        return redirect(url_for('dashboard'))
    rpath = report_path(slug)
    if not os.path.exists(rpath):
        flash('No report yet.', 'error')
        return redirect(url_for('program_detail', slug=slug))
    try:
        from openpyxl import load_workbook
        wb   = load_workbook(rpath, data_only=True)
        data = {sn: [[str(v) if v is not None else '' for v in row]
                     for row in wb[sn].iter_rows(values_only=True)]
                for sn in wb.sheetnames}
        return render_template('edit_report.html', program=p,
                               sheets=data, sheet_names=wb.sheetnames)
    except Exception as e:
        flash(f'Could not open report: {e}', 'error')
        return redirect(url_for('program_detail', slug=slug))

@app.route('/programs/<slug>/save-report', methods=['POST'])
@login_required
def save_report_edits(slug):
    p = get_program(slug, session['user_id'])
    if not p: return jsonify({'success': False, 'error': 'Not found'}), 404
    rpath = report_path(slug)
    if not os.path.exists(rpath):
        return jsonify({'success': False, 'error': 'Report not found'}), 404
    try:
        from openpyxl import load_workbook
        edits = request.json
        wb    = load_workbook(rpath)
        for e in edits:
            sheet = e.get('sheet')
            row   = int(e.get('row')) + 1
            col   = int(e.get('col')) + 1
            val   = e.get('value','')
            if sheet in wb.sheetnames:
                try:    val = float(val) if '.' in str(val) else int(val)
                except: pass
                wb[sheet].cell(row, col, val if val != '' else None)
        wb.save(rpath)
        return jsonify({'success': True, 'saved': len(edits)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/programs/<slug>/upload-nominations', methods=['POST'])
@login_required
def upload_nominations(slug):
    p = get_program(slug, session['user_id'])
    if not p: return jsonify({'success': False, 'error': 'Not found'}), 404
    file = request.files.get('nominations_file')
    if not file: return jsonify({'success': False, 'error': 'No file'})
    rpath = report_path(slug)
    if not os.path.exists(rpath):
        return jsonify({'success': False, 'error': 'Process attendance first'})
    try:
        import pandas as pd
        from smart_report import fill_emails
        df = pd.read_csv(file) if file.filename.endswith('.csv') else pd.read_excel(file)
        cols      = [c.lower().strip() for c in df.columns]
        name_col  = next((df.columns[i] for i,c in enumerate(cols) if 'name'  in c), df.columns[0])
        email_col = next((df.columns[i] for i,c in enumerate(cols) if 'email' in c or 'mail' in c), df.columns[1])
        from utils import normalise_name
        email_map = {normalise_name(str(r[name_col])).lower().strip(): str(r[email_col]).strip()
                     for _,r in df.iterrows() if pd.notna(r[name_col]) and pd.notna(r[email_col])}
        filled = fill_emails(rpath, email_map)
        return jsonify({'success': True, 'filled': filled, 'message': f'{filled} email(s) added'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/programs/<slug>/sessions-json')
@login_required
def sessions_json(slug):
    log = load_log(slug)
    return jsonify([{'date': r['date'], 'title': r['title']} for r in log.get('runs',[])])

@app.route('/admin')
@admin_required
def admin_panel():
    users = get_db().execute('SELECT * FROM users ORDER BY created_at DESC').fetchall()
    return render_template('admin.html', users=users)

@app.route('/admin/users/new', methods=['POST'])
@admin_required
def admin_create_user():
    email    = request.form.get('email','').strip().lower()
    name     = request.form.get('name','').strip()
    password = request.form.get('password','').strip()
    is_admin = 1 if request.form.get('is_admin') else 0
    if not email or not name or not password:
        flash('All fields required.', 'error')
        return redirect(url_for('admin_panel'))
    db = get_db()
    try:
        db.execute('INSERT INTO users (email,name,password,is_admin) VALUES (?,?,?,?)',
                   (email, name, generate_password_hash(password), is_admin))
        db.commit()
        flash(f'Account created for {email}', 'success')
    except sqlite3.IntegrityError:
        flash(f'{email} already exists.', 'error')
    return redirect(url_for('admin_panel'))

@app.route('/admin/users/<int:user_id>/toggle-admin', methods=['POST'])
@admin_required
def toggle_admin(user_id):
    if user_id == session['user_id']:
        flash("Can't change your own admin status.", 'error')
        return redirect(url_for('admin_panel'))
    db   = get_db()
    user = db.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()
    if user:
        db.execute('UPDATE users SET is_admin=? WHERE id=?',
                   (0 if user['is_admin'] else 1, user_id))
        db.commit()
        flash(f'Updated {user["email"]}', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@admin_required
def reset_password(user_id):
    new_pw = request.form.get('new_password','').strip()
    if not new_pw:
        flash('Password cannot be empty.', 'error')
        return redirect(url_for('admin_panel'))
    db = get_db()
    db.execute('UPDATE users SET password=? WHERE id=?',
               (generate_password_hash(new_pw), user_id))
    db.commit()
    flash('Password reset.', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def delete_user(user_id):
    if user_id == session['user_id']:
        flash("Can't delete yourself.", 'error')
        return redirect(url_for('admin_panel'))
    db = get_db()
    db.execute('DELETE FROM program_access WHERE user_id=?', (user_id,))
    db.execute('DELETE FROM users WHERE id=?', (user_id,))
    db.commit()
    flash('User deleted.', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin/users/<int:user_id>/generate-reset', methods=['POST'])
@admin_required
def generate_reset_link(user_id):
    import secrets
    db   = get_db()
    user = db.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()
    if not user:
        flash('User not found.', 'error')
        return redirect(url_for('admin_panel'))
    token = secrets.token_urlsafe(32)
    rf    = os.path.join(DATA_DIR, 'reset_requests.json')
    resets = {}
    if os.path.exists(rf):
        with open(rf) as f:
            try: resets = json.load(f)
            except: pass
    resets[user['email']] = {'token': token, 'created_at': datetime.now().isoformat()}
    with open(rf,'w') as f: json.dump(resets, f)
    url = f"{request.host_url.rstrip('/')}/forgot-password?email={user['email']}&token={token}"
    flash(f'Reset link for {user["email"]}: {url}', 'info')
    return redirect(url_for('admin_panel'))

@app.route('/profile/change-password', methods=['POST'])
@login_required
def change_password():
    current = request.form.get('current_password','').strip()
    new_pw  = request.form.get('new_password','').strip()
    confirm = request.form.get('confirm_password','').strip()
    db      = get_db()
    user    = db.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not check_password_hash(user['password'], current):
        flash('Current password incorrect.', 'error')
    elif new_pw != confirm:
        flash('Passwords do not match.', 'error')
    elif len(new_pw) < 6:
        flash('Must be at least 6 characters.', 'error')
    else:
        db.execute('UPDATE users SET password=? WHERE id=?',
                   (generate_password_hash(new_pw), session['user_id']))
        db.commit()
        flash('Password changed.', 'success')
    return redirect(url_for('dashboard'))



@app.route('/programs/<slug>/paste-feedback', methods=['GET', 'POST'])
@login_required
def paste_feedback(slug):
    p = get_program(slug, session['user_id'])
    if not p:
        flash('Not found.', 'error')
        return redirect(url_for('dashboard'))
    if not os.path.exists(report_path(slug)):
        flash('Process attendance first.', 'error')
        return redirect(url_for('program_detail', slug=slug))

    settings     = json.loads(p['settings'])
    mentor_names = settings.get('mentor_names', [])
    log          = load_log(slug)
    sessions     = [{'date': r['date'], 'title': r['title']}
                    for r in log.get('runs', [])]

    if request.method == 'POST':
        module_name  = request.form.get('module_name', '').strip()
        mentor_name  = request.form.get('mentor_name', '').strip()
        pasted_text  = request.form.get('feedback_text', '').strip()

        if not module_name:
            flash('Module name is required.', 'error')
            return render_template('paste_feedback.html', program=p,
                                   mentor_names=mentor_names, sessions=sessions,
                                   saved_text=pasted_text, saved_mentor=mentor_name)
        if not pasted_text:
            flash('Please paste your feedback data.', 'error')
            return render_template('paste_feedback.html', program=p,
                                   mentor_names=mentor_names, sessions=sessions,
                                   saved_text='', saved_mentor=mentor_name)

        try:
            result = process_feedback_paste(
                pasted_text  = pasted_text,
                module_name  = module_name,
                mentor_name  = mentor_name,
                log_path     = log_path(slug),
                report_path  = report_path(slug),
            )
            if result['success']:
                flash(f"{result['new_rows']} feedback rows added"
                      + (f", {result['skipped_rows']} already existed."
                         if result['skipped_rows'] else '.'), 'success')
                return redirect(url_for('program_detail', slug=slug))
            else:
                flash(f"Error: {result['error']}", 'error')
        except Exception as e:
            import traceback
            flash(f"Error: {str(e)} — {traceback.format_exc()[-300:]}", 'error')

    return render_template('paste_feedback.html', program=p,
                           mentor_names=mentor_names, sessions=sessions,
                           saved_text='', saved_mentor='')


@app.route('/programs/<slug>/remove-participant', methods=['POST'])
@login_required
def remove_participant_route(slug):
    p = get_program(slug, session['user_id'])
    if not p:
        return jsonify({'success': False, 'error': 'Not found'}), 404

    name = request.form.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Name required'})

    rpath = report_path(slug)
    if not os.path.exists(rpath):
        return jsonify({'success': False, 'error': 'No report found'})

    try:
        from smart_report import remove_participant
        removed = remove_participant(rpath, name)

        if removed:
            # Also add to exclude_names in settings so they're skipped in future
            settings = json.loads(p['settings'])
            excl = settings.get('exclude_names', [])
            if name not in excl:
                excl.append(name)
                settings['exclude_names'] = excl
                db = get_db()
                db.execute('UPDATE programs SET settings=? WHERE slug=?',
                           (json.dumps(settings), slug))
                db.commit()

            return jsonify({'success': True,
                            'message': f'{name} removed from report and added to exclude list'})
        else:
            return jsonify({'success': False,
                            'error': f'{name} not found in report'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/programs/<slug>/toggle-complete', methods=['POST'])
@login_required
def toggle_complete(slug):
    p = get_program(slug, session['user_id'])
    if not p:
        flash('Not found.', 'error')
        return redirect(url_for('dashboard'))
    log = load_log(slug)
    log['completed'] = not log.get('completed', False)
    save_log(slug, log)
    status = 'Completed' if log['completed'] else 'In Progress'
    flash(f'Program marked as {status}.', 'success')
    return redirect(url_for('program_detail', slug=slug))

# ── Startup ───────────────────────────────────────────────────────────────────

with app.app_context():
    init_db()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
