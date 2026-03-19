"""
process_feedback_paste.py
Processes feedback pasted directly as text (from CSV copy-paste).
No file upload needed.

Input format - user pastes CSV content, we parse it flexibly:
- Tab separated (from Excel copy)
- Comma separated (from CSV)
- One response per line

We only need: Participant Name + responses
Module name and Mentor name are provided manually.
"""
import os, sys, json
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from openpyxl import load_workbook


def process_feedback_paste(pasted_text, module_name, mentor_name,
                            log_path, report_path):
    """
    Parse pasted feedback text and append to report.
    Returns {'success': True, 'new_rows': N, 'skipped_rows': N}
    """
    if not pasted_text or not pasted_text.strip():
        return {'success': False, 'error': 'No text pasted.'}
    if not module_name or not module_name.strip():
        return {'success': False, 'error': 'Module name is required.'}

    lines = [l for l in pasted_text.strip().splitlines() if l.strip()]
    if not lines:
        return {'success': False, 'error': 'No data found in pasted text.'}

    # Detect separator
    first_line = lines[0]
    sep = '\t' if '\t' in first_line else ','

    # Parse header row (first line)
    headers = [h.strip().strip('"') for h in first_line.split(sep)]

    # Find key columns
    def find_col(candidates):
        for cand in candidates:
            for i, h in enumerate(headers):
                if cand.lower() in h.lower():
                    return i
        return None

    participant_col = find_col(['participant', 'name', 'respondent', 'your name'])
    rating_col      = find_col(['rating', 'rate', 'score', 'stars'])
    takeaway_col    = find_col(['takeaway', 'key learning', 'what did you'])
    specific_col    = find_col(['specific', 'feedback for', 'delivery', 'mentor'])
    other_col       = find_col(['other', 'additional', 'anything'])
    date_col        = find_col(['start time', 'timestamp', 'date', 'completion'])

    if participant_col is None:
        # No header — treat all lines as participant names only
        participant_col = 0

    # Load workbook
    if not os.path.exists(report_path):
        return {'success': False, 'error': 'Report not found. Upload attendance first.'}

    wb = load_workbook(report_path)
    if 'Feedback' not in wb.sheetnames:
        return {'success': False, 'error': 'No Feedback sheet in report.'}

    ws = wb['Feedback']

    # Find next empty row + max Sno
    next_row = 2
    while ws.cell(next_row, 1).value is not None:
        next_row += 1

    max_sno = 0
    for row in range(2, next_row):
        v = ws.cell(row, 1).value
        try:
            if v: max_sno = max(max_sno, int(str(v)))
        except: pass

    # Load existing keys for dedup
    log = {}
    if os.path.exists(log_path):
        with open(log_path) as f:
            try: log = json.load(f)
            except: pass
    done_keys  = set(log.get('processed_feedback_keys', []))
    new_count  = 0
    skip_count = 0
    timestamp  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def get_cell(row_cells, idx):
        if idx is None or idx >= len(row_cells):
            return ''
        return row_cells[idx].strip().strip('"')

    # Skip header line, process data lines
    data_lines = lines[1:] if len(lines) > 1 else lines

    for line in data_lines:
        cells = line.split(sep)
        participant = get_cell(cells, participant_col)
        if not participant:
            continue

        rating      = get_cell(cells, rating_col)
        takeaways   = get_cell(cells, takeaway_col)
        specific    = get_cell(cells, specific_col)
        other       = get_cell(cells, other_col)
        ts          = get_cell(cells, date_col) or timestamp

        key = f"{ts}|{participant.lower()}|{module_name.lower()}"
        if key in done_keys:
            skip_count += 1
            continue

        max_sno += 1
        ws.cell(next_row, 1, max_sno)
        ws.cell(next_row, 2, ts)
        ws.cell(next_row, 3, module_name)
        ws.cell(next_row, 4, mentor_name)
        ws.cell(next_row, 5, participant)
        ws.cell(next_row, 6, takeaways)
        try:    ws.cell(next_row, 7, float(rating))
        except: ws.cell(next_row, 7, rating)
        ws.cell(next_row, 8, specific)
        ws.cell(next_row, 9, other)

        done_keys.add(key)
        next_row  += 1
        new_count += 1

    wb.save(report_path)

    log['processed_feedback_keys'] = list(done_keys)
    log.setdefault('feedback_runs', []).append({
        'method':       'paste',
        'module':       module_name,
        'mentor':       mentor_name,
        'new_rows':     new_count,
        'skipped':      skip_count,
        'processed_at': datetime.now().isoformat()
    })
    os.makedirs(os.path.dirname(log_path) or '.', exist_ok=True)
    with open(log_path, 'w') as f:
        json.dump(log, f, indent=2, default=str)

    return {'success': True, 'new_rows': new_count, 'skipped_rows': skip_count}
