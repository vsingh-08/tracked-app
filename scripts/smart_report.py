"""
smart_report.py
Builds and updates the attendance report to match the template exactly.
No template needed — everything is built from the CSV automatically.

Sheet structure (matches template):
  Consolidated Report  — Attendance + Attentiveness per session
  Overall Attendance   — Yes/No per session + totals + %
  Login                — Join/Leave times per session
  Feedback             — Filled from Forms export
"""
import os, sys, re, copy
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from utils import parse_teams_csv, normalise_name

# ── Exact colours from template ───────────────────────────────────────────────
C_DARK_BLUE   = 'FF002060'  # Fixed col headers (Sl No, Names, Email)
C_MED_BLUE    = 'FF4472C4'  # Session headers in Consolidated + Overall
C_LIGHT_BLUE  = 'FFD9E1F2'  # Sub-headers (Attendance / Attentiveness)
C_OVR_FIXED   = 'FF002060'  # Overall Attendance fixed col headers
C_LOGIN_HDR   = 'FF61CBF3'  # Login session header bg
C_LOGIN_SUB   = 'FFF7C7AC'  # Login sub-header bg (Name, First Join...)
C_WHITE_FONT  = 'FFFFFFFF'
C_BLACK_FONT  = 'FF000000'

# ── Layout constants (match template exactly) ─────────────────────────────────
# Consolidated Report:
#   C1=Sl No  C2=Names  C3=Email  C4=spacer
#   Session N: base_col = 5 + (N-1)*3
#              base_col   = Attendance (merged header spans base_col:base_col+1)
#              base_col+1 = Attentiveness
#              base_col+2 = spacer
CONS_FIXED     = 4          # cols before first session
CONS_PER_SES   = 3          # Attendance + Attentiveness + spacer

# Overall Attendance:
#   C1=S.No C2=Name C3=Conducted C4=Attended C5=% C6=spacer
#   Session N: col = 7 + (N-1)
OVR_FIXED      = 7
OVR_PER_SES    = 1

# Login:
#   Session N block starts at col: 1 + (N-1)*5
#   Row 1+2 merged: session header (4 cols wide)
#   Row 3: Name | First Join | Last Leave | In-Meeting Duration
#   Row 4+: data
#   Col 5 of each block = spacer
LOGIN_PER_SES  = 5
LOGIN_DATA_ROW = 4


def _f(bold=False, size=11, rgb=None):
    kw = dict(bold=bold, size=size)
    if rgb: kw['color'] = rgb
    return Font(**kw)

def _fill(rgb):
    return PatternFill(fill_type='solid', fgColor=rgb)

def _al(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    s = Side(style='thin', color='FFD3D3D3')
    return Border(left=s, right=s, top=s, bottom=s)

def _style(cell, bold=False, font_rgb=None, fill_rgb=None, h='center', wrap=False):
    cell.font      = _f(bold=bold, rgb=font_rgb)
    cell.alignment = _al(h=h, wrap=wrap)
    cell.border    = _border()
    if fill_rgb:
        cell.fill  = _fill(fill_rgb)

# ── Column helpers ────────────────────────────────────────────────────────────

def cons_att_col(si):   return CONS_FIXED + 1 + si * CONS_PER_SES
def cons_atten_col(si): return CONS_FIXED + 2 + si * CONS_PER_SES
def ovr_col(si):        return OVR_FIXED  + si * OVR_PER_SES
def login_base(si):     return 1          + si * LOGIN_PER_SES


# ── Session detection ─────────────────────────────────────────────────────────

def _date_variants(date_str):
    v = [date_str]
    p = date_str.split('-')
    if len(p) == 3:
        d,m,y = p
        v += [f"{d}/{m}/{y}", f"{int(d)}-{m}-{y}", f"{int(d)}/{m}/{y}",
              f"{d}-{m}-{y[2:]}", f"{int(d)}-{m}-{y[2:]}"]
    return v

def find_session_idx(wb, date_str):
    ws = wb['Consolidated Report']
    dv = _date_variants(date_str)
    for si in range(50):
        col = cons_att_col(si)
        if col > ws.max_column + CONS_PER_SES:
            break
        # Check merged header cell (one col before att_col)
        for check_col in [col - 1, col]:
            val = ws.cell(1, check_col).value
            if not val: continue
            s = str(val).replace('\n',' ').replace('\xa0',' ')
            for dv_ in dv:
                if dv_ in s:
                    return si
    return None

def count_sessions(wb):
    ws = wb['Consolidated Report']
    for si in range(50):
        col = cons_att_col(si)
        v   = ws.cell(2, col).value
        if not v or 'attendance' not in str(v).lower():
            return si
    return 0

def get_roster(wb):
    ws = wb['Consolidated Report']
    names = []
    for row in range(3, ws.max_row + 1):
        v = ws.cell(row, 2).value
        if v: names.append(str(v))
    return names

def find_name_row(ws, name, name_col=2, start=3):
    norm = normalise_name(str(name)).lower().strip()
    for row in range(start, ws.max_row + 1):
        v = ws.cell(row, name_col).value
        if v and normalise_name(str(v)).lower().strip() == norm:
            return row
    return None


# ── Create brand-new workbook ─────────────────────────────────────────────────

def create_workbook(names):
    wb  = Workbook()
    ws  = wb.active;  ws.title = 'Consolidated Report'
    ws2 = wb.create_sheet('Overall Attendance')
    ws3 = wb.create_sheet('Login')
    ws4 = wb.create_sheet('Feedback')

    _init_consolidated(ws,  names)
    _init_overall(ws2, names)
    _init_login(ws3)
    _init_feedback(ws4)
    return wb


def _init_consolidated(ws, names):
    # Fixed headers
    for col, val in [(1,'Sl  No'),(2,'Names '),(3,'Email')]:
        c = ws.cell(1, col, val)
        _style(c, bold=True, font_rgb=C_WHITE_FONT, fill_rgb=C_DARK_BLUE, wrap=True)

    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')

    # Participant rows
    for i, name in enumerate(names):
        row = 3 + i
        for col, val in [(1, i+1),(2, name)]:
            c = ws.cell(row, col, val)
            _style(c)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 3
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 16


def _init_overall(ws, names):
    hdrs = [(1,'S. No'),(2,'Name'),(3,'Total sessions Conducted'),
            (4,'Total sessions attended'),(5,'% of attended sessions')]
    for col, val in hdrs:
        c = ws.cell(1, col, val)
        _style(c, bold=True, font_rgb=C_WHITE_FONT, fill_rgb=C_OVR_FIXED, wrap=True)

    # "Overall" label under % col
    c = ws.cell(2, 5, 'Overall')
    _style(c, bold=True, fill_rgb=C_LIGHT_BLUE)

    for i, name in enumerate(names):
        row = 3 + i
        ws.cell(row, 1, i+1)
        ws.cell(row, 2, name)
        # Write 0 initially — updated by _recalc_overall() after each session
        ws.cell(row, 3, 0)
        ws.cell(row, 4, 0)
        ws.cell(row, 5, 0)
        ws.cell(row, 5).number_format = '0%'
        for col in [1,2,3,4,5]:
            _style(ws.cell(row, col))

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 3
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 16


def _init_login(ws):
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 5
    ws.row_dimensions[3].height = 16


def _init_feedback(ws):
    headers = ['Sno','Date','Module Name','Faculty','Participant Name',
               "Please share your key takeaways from today's session",
               'My rating',
               'Please let us know if you have any specific feedback for the module / delivery method / mentor?',
               'Any other feedback']
    for i, h in enumerate(headers):
        c = ws.cell(1, i+1, h)
        _style(c, bold=True, font_rgb=C_WHITE_FONT, fill_rgb=C_DARK_BLUE, wrap=True)

    widths = [6, 22, 30, 20, 20, 50, 12, 50, 30]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w
    ws.row_dimensions[1].height = 30


# ── Add session columns ───────────────────────────────────────────────────────

def add_session(wb, session_name, session_date, session_idx):
    header = f"{session_name}\n ({session_date})"

    # ── Consolidated Report ───────────────────────────────────────────────────
    ws  = wb['Consolidated Report']
    ac  = cons_att_col(session_idx)      # Attendance col
    avc = cons_atten_col(session_idx)    # Attentiveness col
    bc  = ac - 1                          # Base col for merge (ac is base in template)

    # Merge header over Attendance + Attentiveness cols
    ws.merge_cells(start_row=1, start_column=ac,
                   end_row=1,   end_column=avc)
    c = ws.cell(1, ac, header)
    _style(c, bold=True, font_rgb=C_WHITE_FONT, fill_rgb=C_MED_BLUE, wrap=True)

    # Sub-headers row 2
    for col, val in [(ac,'Attendance'),(avc,'Attentiveness')]:
        c = ws.cell(2, col, val)
        _style(c, bold=True, fill_rgb=C_LIGHT_BLUE)

    # Spacer col
    ws.column_dimensions[get_column_letter(avc+1)].width = 3
    ws.column_dimensions[get_column_letter(ac)].width    = 14
    ws.column_dimensions[get_column_letter(avc)].width   = 16

    # ── Overall Attendance ────────────────────────────────────────────────────
    ws2 = wb['Overall Attendance']
    oc  = ovr_col(session_idx)

    c = ws2.cell(1, oc, header)
    _style(c, bold=True, font_rgb=C_WHITE_FONT, fill_rgb=C_MED_BLUE, wrap=True)

    c = ws2.cell(2, oc, 'Attendance')
    _style(c, bold=True, fill_rgb=C_LIGHT_BLUE)

    ws2.column_dimensions[get_column_letter(oc)].width = 14

    # ── Login ─────────────────────────────────────────────────────────────────
    ws3  = wb['Login']
    lb   = login_base(session_idx)

    # Merge header over 4 cols (rows 1-2), exactly like template
    ws3.merge_cells(start_row=1, start_column=lb,
                    end_row=2,   end_column=lb+3)
    c = ws3.cell(1, lb, header)
    c.font      = _f(bold=True, rgb=C_BLACK_FONT)
    c.fill      = _fill(C_LOGIN_HDR)
    c.alignment = _al(h='center', wrap=True)

    # Sub-headers row 3
    for sc, sub in enumerate(['Name','First Join','Last Leave','In-Meeting Duration']):
        col = lb + sc
        c   = ws3.cell(3, col, sub)
        c.font      = _f(bold=True)
        c.fill      = _fill(C_LOGIN_SUB)
        c.alignment = _al(h='center')
        ws3.column_dimensions[get_column_letter(col)].width = 20 if sc==0 else 24

    # Spacer col
    ws3.column_dimensions[get_column_letter(lb+4)].width = 3


# ── Fill session data ─────────────────────────────────────────────────────────

def fill_data(wb, participants_data, session_idx):
    ws  = wb['Consolidated Report']
    ws2 = wb['Overall Attendance']
    ws3 = wb['Login']

    ac  = cons_att_col(session_idx)
    avc = cons_atten_col(session_idx)
    oc  = ovr_col(session_idx)
    lb  = login_base(session_idx)

    # Consolidated + Overall
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row, 2).value
        if not name: continue
        norm = normalise_name(str(name)).lower().strip()

        match = None
        for pname, pdata in participants_data.items():
            if normalise_name(pname).lower().strip() == norm:
                match = pdata
                break

        att_val  = match['attendance']    if match else 'No'
        atten_val= match['attentiveness'] if match else 0

        c1 = ws.cell(row, ac,  att_val)
        c2 = ws.cell(row, avc, atten_val)
        c2.number_format = '0%'
        _style(c1)
        _style(c2)

        row2 = find_name_row(ws2, str(name))
        if row2:
            c = ws2.cell(row2, oc, att_val)
            _style(c)

    # Login — present participants sorted by join time
    present = sorted(
        [(n,d) for n,d in participants_data.items() if d['attendance']=='Yes'],
        key=lambda x: x[1]['first_join']
    )
    for i, (name, d) in enumerate(present):
        lrow = LOGIN_DATA_ROW + i
        ws3.cell(lrow, lb,   name)
        ws3.cell(lrow, lb+1, d['first_join'])
        ws3.cell(lrow, lb+2, d['last_leave'])
        ws3.cell(lrow, lb+3, d['duration_str'])


def add_participants(wb, new_names):
    for sheet_name in ['Consolidated Report', 'Overall Attendance']:
        ws = wb[sheet_name]
        next_row = 3
        while ws.cell(next_row, 2).value:
            next_row += 1
        start_i = next_row - 3

        for i, name in enumerate(new_names):
            row = next_row + i
            ws.cell(row, 1, start_i + i + 1)
            ws.cell(row, 2, name)
            _style(ws.cell(row, 1))
            _style(ws.cell(row, 2))

            if sheet_name == 'Overall Attendance':
                ws.cell(row, 3).value = '=COUNTIF($G$2:$XFD$2,"Attendance")'
                ws.cell(row, 4).value = f'=COUNTIF(G{row}:XFD{row},"Yes")'
                ws.cell(row, 5).value = f'=IFERROR(D{row}/C{row},0)'
                ws.cell(row, 5).number_format = '0%'
                for col in [1,2,3,4,5]:
                    _style(ws.cell(row, col))


# ── Main entry point ──────────────────────────────────────────────────────────

def _backfill_absent(wb, new_names):
    """Fill previous sessions with No/0 for newly added participants."""
    ws  = wb['Consolidated Report']
    ws2 = wb['Overall Attendance']

    # Find all existing session columns
    existing_sessions = []
    for si in range(count_sessions(wb)):
        att_col  = cons_att_col(si)
        atten_col= cons_atten_col(si)
        existing_sessions.append((att_col, atten_col, ovr_col(si)))

    if not existing_sessions:
        return

    for name in new_names:
        # Find the row just added for this participant
        row = find_name_row(ws, name)
        if not row:
            continue
        for att_col, atten_col, o_col in existing_sessions:
            # Only fill if currently empty
            if not ws.cell(row, att_col).value:
                ws.cell(row, att_col,   'No')
                ws.cell(row, atten_col, 0)
                ws.cell(row, atten_col).number_format = '0%'
                _style(ws.cell(row, att_col))
                _style(ws.cell(row, atten_col))

        # Overall Attendance sheet
        row2 = find_name_row(ws2, name)
        if row2:
            for _, _, o_col in existing_sessions:
                if not ws2.cell(row2, o_col).value:
                    ws2.cell(row2, o_col, 'No')
                    _style(ws2.cell(row2, o_col))


def remove_participant(report_path, name):
    """
    Remove a participant from all sheets in the report.
    Returns True if found and removed, False if not found.
    """
    wb  = load_workbook(report_path)
    ws  = wb['Consolidated Report']
    ws2 = wb['Overall Attendance']
    ws3 = wb['Login']

    removed = False

    for sheet, ws_obj in [('Consolidated Report', ws),
                          ('Overall Attendance', ws2)]:
        row = find_name_row(ws_obj, name)
        if row:
            ws_obj.delete_rows(row)
            # Renumber S.No column
            for r in range(3, ws_obj.max_row + 1):
                if ws_obj.cell(r, 2).value:
                    ws_obj.cell(r, 1, r - 2)
            removed = True

    wb.save(report_path)
    return removed


def _recalc_overall(wb):
    """Recalculate Total Conducted, Attended, % with actual values."""
    ws = wb['Overall Attendance']
    session_cols = [col for col in range(1, ws.max_column + 1)
                    if str(ws.cell(2, col).value or '').strip() == 'Attendance']
    session_count = len(session_cols)
    if not session_count:
        return
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row, 2).value
        if not name:
            continue
        yes_count = sum(1 for col in session_cols
                        if str(ws.cell(row, col).value or '').strip() == 'Yes')
        ws.cell(row, 3, session_count)
        ws.cell(row, 4, yes_count)
        pct = round((yes_count / session_count) * 100, 2) if session_count else 0
        ws.cell(row, 5, pct)
        ws.cell(row, 5).number_format = '0.00"%"'
        _style(ws.cell(row, 3))
        _style(ws.cell(row, 4))
        _style(ws.cell(row, 5))


def fill_emails(report_path, email_map):
    """
    Fill email column in Consolidated Report.
    email_map: {normalised_name: email}
    Called after uploading a nominations CSV.
    """
    wb = load_workbook(report_path)
    ws = wb['Consolidated Report']
    filled = 0
    for row in range(3, ws.max_row+1):
        name = ws.cell(row, 2).value
        if not name: continue
        norm = normalise_name(str(name)).lower().strip()
        # Try exact match first, then partial
        email = email_map.get(norm)
        if not email:
            for k, v in email_map.items():
                if k in norm or norm in k:
                    email = v
                    break
        if email:
            ws.cell(row, 3, email)
            filled += 1
    wb.save(report_path)
    return filled


def process_csv(csv_path, report_path,
                exclude_names=None, mentor_names=None,
                threshold=0.50, name_format='auto'):

    exclude_set = {normalise_name(n, name_format).lower()
                   for n in (exclude_names or [])}
    mentor_set  = {normalise_name(n, name_format).lower()
                   for n in (mentor_names  or [])}
    all_excl    = exclude_set | mentor_set

    # ── Parse CSV ─────────────────────────────────────────────────────────────
    session_info, raw_parts, warnings = parse_teams_csv(csv_path)

    if session_info.get('date') == 'Unknown':
        return {'success': False,
                'error': f"Could not read date. Expected: M/DD/YY, H:MM:SS AM/PM"}

    date     = session_info['date']
    duration = session_info['duration_seconds']
    title    = session_info.get('meeting_title', 'Session')

    if not duration:
        return {'success': False, 'error': 'Could not read session duration.'}

    # ── Filter ────────────────────────────────────────────────────────────────
    filtered = []
    for p in raw_parts:
        if p['role'].lower() == 'organizer': continue
        norm = normalise_name(p['name'], name_format).lower().strip()
        if norm in all_excl or p['name'].lower() in all_excl: continue
        filtered.append(p)

    # ── Aggregate rejoins ─────────────────────────────────────────────────────
    agg = {}
    for p in filtered:
        key     = normalise_name(p['name'], name_format).lower().strip()
        display = normalise_name(p['name'], name_format)
        if key in agg:
            agg[key]['duration_seconds'] += p['duration_seconds']
        else:
            agg[key] = {**p, 'display': display}

    # ── Compute attendance ────────────────────────────────────────────────────
    pdata = {}
    for key, p in agg.items():
        att = min(round(p['duration_seconds'] / duration, 4), 1.0)
        pdata[p['display']] = {
            'attendance':    'Yes' if att >= threshold else 'No',
            'attentiveness': att,
            'first_join':    p['first_join'],
            'last_leave':    p['last_leave'],
            'duration_str':  p['duration_str'],
        }

    present = sum(1 for d in pdata.values() if d['attendance']=='Yes')
    absent  = [n for n,d in pdata.items()   if d['attendance']=='No']

    # ── Load or create workbook ───────────────────────────────────────────────
    if os.path.exists(report_path):
        wb = load_workbook(report_path)

        # Check already processed
        if find_session_idx(wb, date) is not None:
            return {'success': False,
                    'error': f'Session {date} already in report. '
                             'Move it to processed/ and delete it from the report to reprocess.'}

        # Add any new participants
        existing = {normalise_name(n).lower() for n in get_roster(wb)}
        new_names = [n for n in pdata
                     if normalise_name(n).lower() not in existing]
        if new_names:
            print(f"  Adding {len(new_names)} new participant(s): {new_names}")
            add_participants(wb, new_names)
            # Backfill previous sessions with "No" for new participants
            _backfill_absent(wb, new_names)

        si = count_sessions(wb)
    else:
        print(f"  Creating new report...")
        wb = create_workbook(sorted(pdata.keys()))
        si = 0

    # ── Add session + fill data ───────────────────────────────────────────────
    print(f"  Session {si+1}: {title} ({date})")
    add_session(wb, title, date, si)
    fill_data(wb, pdata, si)

    # Recalculate overall attendance totals with actual values
    _recalc_overall(wb)

    os.makedirs(os.path.dirname(report_path) or '.', exist_ok=True)
    wb.save(report_path)

    return {
        'success': True, 'date': date, 'title': title,
        'session_num': si+1, 'present': present, 'absent': absent,
        'participants': list(pdata.keys())
    }
